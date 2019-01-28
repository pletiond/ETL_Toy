import logging
import configparser
from openpyxl import Workbook


class Write_To_Excel:

    def __init__(self, name, data_target_name):
        self.data = None
        self.data_target_name = data_target_name
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None

    def process(self):
        self.logger.info(f'Starting new write_to_excel job - {self.name}!')
        file, sheet = self._get_file_and_sheet()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        self._write_header(ws)
        self._write_data(ws)
        wb.save(file)

        self.logger.info(
            f'Write_to_excel job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _get_file_and_sheet(self):
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            raise Exception('Invalid data target')
        if not 'file' in config[self.data_target_name]:
            raise Exception('Invalid data target type')
        if not 'sheet' in config[self.data_target_name]:
            raise Exception('Invalid data target type')
        return config[self.data_target_name]['file'], config[self.data_target_name]['sheet']

    def _write_header(self, ws):
        for i in range(len(self.data.columns_names)):
            _ = ws.cell(1, i + 1, self.data.columns_names[i])

    def _write_data(self, ws):
        for row in range(len(self.data.data)):
            for column in range(len(self.data.columns_names)):
                _ = ws.cell(row + 2, column + 1, self.data.data[row][column])
