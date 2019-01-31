import configparser
import logging

import numpy
import psycopg2
from openpyxl import load_workbook, Workbook


class Dummy:
    """
    Dummy class for testing
    """

    def __init__(self, name):
        """
        :param name: Step name
        :type name: str
        """
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data = None

    def process(self):
        self.logger.info(f'Starting new dummy job - {self.name}!')
        self.logger.info(
            f'Dummy job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')


class Filter_Columns:
    """
    This step filter column by whitelist or blacklist
    """

    def __init__(self, name):
        """
        :param name: Step name
        :type name: str
        """
        self.name = name
        self.allowed = None
        self.remove = None
        self.data = None
        self.logger = logging.getLogger(__name__)

    def set_columns(self, names):
        """
        :param names: Column names
        :type names: list
        """
        self.allowed = names

    def remove_columns(self, names):
        """
        :param names: columns to remove
        :type names: list
        """
        self.remove = names

    def process(self):
        self.logger.info(f'Starting new filter_columns job - {self.name}!')

        if self.allowed:
            self._select_columns()
        elif self.remove:
            self._only_remove()

        self.logger.info(
            f'Filter_columns job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _only_remove(self):
        """
        Remove selected columns
        :return: False if column doesnt exist
        """
        for column in self.remove:
            res = self.data.remove_column(column)
            if not res:
                self.logger.info(f'{self.name} - Column doesnt exist!')
                raise Exception('{self.name} - Column doesnt exist!')

    def _select_columns(self):
        """
        Select columns by whitelist, remove others
        """
        all_columns = list(self.data.columns_names)
        for column in all_columns:
            if column in self.allowed:
                continue
            self.data.remove_column(column)


class Mapping:
    """
    Using rules map values in selected column to new ones
    """

    def __init__(self, name):
        """
        :param name: Step name
        :type name: str
        """
        self.name = name
        self.data = None
        self.logger = logging.getLogger(__name__)
        self.rules = dict()
        self.transpone = None
        self.transpone_final = None

    def new_mapping(self, column, old, new):
        """
        Add new mapping rule

        :param column: Selected column
        :param old: Old value
        :param new: New value
        """
        if not column in self.rules:
            self.rules[column] = {}
        self.rules[column][old] = new

    def process(self):
        self.logger.info(f'Starting new mapping job - {self.name}!')

        self.transpone = self.data.data.T
        self.transpone_final = self.data.data.T

        for column, rules in self.rules.items():
            self._apply_rules(column, rules)

        self.data.data = self.transpone_final.T
        self.logger.info(
            f'mapping job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _apply_rules(self, column, rules):
        """
        Apply rules in selected column
        :param column: Selected column
        :type column: str
        :param rules: Rules
        :type rules: dict
        """
        for i in range(len(self.data.columns_names)):
            if not column == self.data.columns_names[i]:
                continue

            orig = self.transpone[i]

            for old, new in rules.items():
                for k in range(len(self.transpone[i])):
                    if str(orig[k]) == str(old):
                        orig[k] = new

            self.transpone_final[i] = orig


class Read_csv:
    """
    Read file and load data
    """

    def __init__(self, name, data_target_name, delimiter):
        """
        :param name: Step name
        :type name: str
        :param data_target_name: Name of data target
        :type data_target_name: str
        :param delimiter: CSV delimiter
        :type delimiter: str
        """
        self.data = None
        self.data_target_name = data_target_name
        self.delimiter = delimiter
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.column_cnt = 0
        self.data_targets = None

    def process(self):
        self.logger.info(f'Starting new read_csv job - {self.name}!')

        first = True
        file = self._get_file()
        with open(file, 'r', encoding='utf8') as f:
            for line in f.readlines():
                if first:
                    self._parse_header(line)
                    first = False
                else:
                    self._parse_line(line)

        self.logger.info(
            f'Read_csv job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _parse_header(self, header):
        """
        Parse  header and add column names to data class
        :param header: Header line
        """
        for column in str(header).replace('\n', '').rstrip().split(self.delimiter):
            self.data.add_column_name(column)
            self.column_cnt += 1

    def _parse_line(self, line):
        """
        Parse line with values
        :param line: Value line
        """
        parsed_line = str(line).replace('\n', '').rstrip().split(self.delimiter)
        if not len(parsed_line) == self.column_cnt:
            self.logger.info(f'{self.name} - header and row lenght not match!')
            raise Exception(f'{self.name} - header and row lenght not match!')

        self.data.add_row(parsed_line)

    def _get_file(self):
        """
        Using data target name find path to csv file
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            self.logger.info(f'{self.name} - Invalid data target name')
            raise Exception('Invalid data target name')
        if not 'file' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - Invalid data target type')
            raise Exception('Invalid data target type')
        return config[self.data_target_name]['file']


class Read_excel:
    """
    Read excel file and load data
    """

    def __init__(self, name, data_target_name):
        """
        :param name: Step name
        :param data_target_name: Data target name
        """
        self.data = None
        self.data_target_name = data_target_name
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None
        self.column_cnt = 0

    def process(self):
        """
        Open workbook and iter all rows
        :return:
        """
        self.logger.info(f'Starting new read_excel job - {self.name}!')

        file, sheet = self._get_file_and_sheet()
        wb = load_workbook(filename=file, read_only=True)
        ws = wb[sheet]
        header = True
        for row in ws.rows:
            if header:
                self._parse_header(row)
                header = False
            else:
                self._parse_row(row)

        self.logger.info(
            f'Read_excel job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _get_file_and_sheet(self):
        """
        Using data target name find path to excel file and sheet name
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            self.logger.info(f'{self.name} - Invalid data target name')
            raise Exception('Invalid data target  name')
        if not 'file' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - Invalid data target type, missing "file"')
            raise Exception('Invalid data target type')
        if not 'sheet' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - Invalid data target type, missing "sheet"')
            raise Exception('Invalid data target type')
        return config[self.data_target_name]['file'], config[self.data_target_name]['sheet']

    def _parse_header(self, row):
        for cell in row:
            self.data.add_column_name(cell.value)
            self.column_cnt += 1

    def _parse_row(self, row):
        line = []
        for cell in row:
            line.append(cell.value)
        if not len(line) == self.column_cnt:
            self.logger.info(f'{self.name} - Different row lenght')
            raise Exception('Different row lenght')
        else:
            self.data.add_row(line)


class Read_Postgresql:
    """
    Connect to Postgresql database and load data from selected table
    """

    def __init__(self, name, data_target_name, table, schema='public', columns='*'):
        """
        :param name: Step name
        :param data_target_name: Data target name
        :param table: Target table name
        :param schema: Target schema name
        :param columns: Columns to load
        """
        self.data = None
        self.data_target_name = data_target_name
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None
        self.conn = None
        self.host = None
        self.port = None
        self.user = None
        self.password = None
        self.schema = schema
        self.table = table
        self.columns = columns

    def process(self):
        """
        Connect to database
        """
        self.logger.info(f'Starting new read_postgresql job - {self.name}!')
        if len(self.data.columns_names) > 0:
            self.logger.info(f'{self.name} not empty input data!')
            raise Exception(f'{self.name} not empty input data!')

        self._load_credentials()
        try:
            self.conn = psycopg2.connect(
                f'dbname={self.dbname} user={self.user} port={self.port} host={self.host} password={self.password}')
        except:
            self.logger.info(f'{self.name} - Unable to connect to database.')
            raise Exception('Unable to connect to database.')

        self._load_data()

        self.logger.info(
            f'Read_postgresql job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _load_data(self):
        """
        Run select query and get all rows
        :return:
        """
        cur = self.conn.cursor()
        cur.execute(f'SELECT {self.columns} FROM {self.schema}.{self.table}')
        colnames = [desc[0] for desc in cur.description]
        row = cur.fetchone()
        for colname in colnames:
            self.data.add_column_name(colname, '')
        while row is not None:
            self.data.add_row(row)
            row = cur.fetchone()

        cur.close()
        self.conn.close()

    def _load_credentials(self):
        """
        Using data target name load host, dbname, port and user
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            self.logger.info(f'{self.name} - not valid data target name!')
            raise Exception('Invalid data target name')
        if not 'dbname' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing dbname value!')
            raise Exception('Invalid data target type')
        if not 'host' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing host value!')
            raise Exception('Invalid data target type')
        if not 'port' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing port value!')
            raise Exception('Invalid data target type')
        if not 'user' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing user value!')
            raise Exception('Invalid data target type')
        if not 'password' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing password value!')
            raise Exception('Invalid data target type')

        self.dbname = config[self.data_target_name]['dbname']
        self.host = config[self.data_target_name]['host']
        self.port = config[self.data_target_name]['port']
        self.user = config[self.data_target_name]['user']
        self.password = config[self.data_target_name]['password']


class Sort_Data():
    """
    Sort dataset by selected column
    """

    def __init__(self, name, column):
        """
        :param name: Step name
        :param column: Sorting column
        """
        self.data = None
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None
        self.sorting_column = column

    def process(self):
        self.logger.info(f'Starting new sorting job - {self.name}!')

        if not self.sorting_column in self.data.columns_names:
            self.logger.info(f'{self.name} column doesnt exist!')
            raise Exception(f'{self.name} column doesnt exist!')

        self.data.sort_by_column(self.sorting_column)

        self.logger.info(
            f'Sort_data job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')


class String_operations:
    """
    Step for string operations - string upper/lower/capitalize
    """

    def __init__(self, name):
        """
        :param name: step name
        """
        self.name = name
        self.data = None
        self.logger = logging.getLogger(__name__)
        self.transpone = None
        self.transpone_final = None
        self.upper = []
        self.lower = []
        self.capitalize = []

    def do_upper(self, columns):
        """Select columns for upper
        :param columns: Selected  columns
        :type columns: list
        """
        self.upper = columns

    def do_lower(self, columns):
        """Select columns for lower
        :param columns: Selected  columns
        :type columns: list
        """
        self.lower = columns

    def do_capitalize(self, columns):
        """Select columns for capitalize
        :param columns: Selected  columns
        :type columns: list
        """
        self.capitalize = columns

    def process(self):
        self.logger.info(f'Starting new string_operations job - {self.name}!')
        self._check_columns()
        self.transpone = self.data.data.T
        self.transpone_final = self.data.data.T

        self._process_lower()
        self._process_upper()
        self._process_capitalize()

        self.data.data = self.transpone_final.T
        self.logger.info(
            f'String_operations job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _check_columns(self):
        for column in self.upper + self.lower + self.capitalize:
            if not column in self.data.columns_names:
                self.logger.info(f'{self.name} - column doesnt exist!')
                raise Exception(f'{self.name} - column doesnt exist!')

    def _process_lower(self):
        for column in self.lower:
            index = self.data.columns_names.index(column)
            self.transpone_final[index] = numpy.char.lower(numpy.array(self.transpone[index]))

    def _process_upper(self):
        for column in self.upper:
            index = self.data.columns_names.index(column)
            self.transpone_final[index] = numpy.char.upper(numpy.array(self.transpone[index]))

    def _process_capitalize(self):
        for column in self.capitalize:
            index = self.data.columns_names.index(column)
            self.transpone_final[index] = numpy.char.capitalize(numpy.array(self.transpone[index]))


class Write_To_CSV:
    """
    Write data to file with CSV format
    """

    def __init__(self, name, data_target_name, delimiter):
        """
        :param name: Step name
        :param data_target_name: Data target name
        :param delimiter: CSV delimiter
        """
        self.name = name
        self.data_target_name = data_target_name
        self.data_targets = None
        self.delimiter = delimiter
        self.data = None
        self.logger = logging.getLogger(__name__)

    def process(self):
        self.logger.info(f'Starting new write_to_csv job - {self.name}!')
        file = self._get_file()
        with open(file, 'w+', encoding='utf8') as f:
            header = str(self.delimiter).join(self.data.columns_names)
            f.write(header + '\n')
            for line in self.data.data:
                f.write(str(self.delimiter).join(line) + '\n')

        self.logger.info(
            f'Write_to_csv job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _get_file(self):
        """
        Load target file from data targets file
        :return: path to file
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            raise Exception('Invalid data target')
        if not 'file' in config[self.data_target_name]:
            raise Exception('Invalid data target type')
        return config[self.data_target_name]['file']


class Write_To_Excel:
    """
    Write data to new Excel workbook
    """

    def __init__(self, name, data_target_name):
        self.data = None
        self.data_target_name = data_target_name
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None

    def process(self):
        """
        Create new workbook and write header and data
        :return:
        """
        self.logger.info(f'Starting new write_to_excel job - {self.name}!')
        file, sheet = self._get_file_and_sheet()

        wb = Workbook()
        ws = wb.active
        ws.title = sheet

        self._write_header(ws)
        self._write_data(ws)
        wb.save(file)

        self.logger.info(
            f'Write_to_excel job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _get_file_and_sheet(self):
        """
        Get file path and sheet from data targets file
        :return: path to file and name of sheet
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            raise Exception('Invalid data target')
        if not 'file' in config[self.data_target_name]:
            raise Exception('Invalid data target type')
        if not 'sheet' in config[self.data_target_name]:
            raise Exception('Invalid data target type')
        return config[self.data_target_name]['file'], config[self.data_target_name]['sheet']

    def _write_header(self, ws):
        for i in range(len(self.data.columns_names)):
            _ = ws.cell(1, i + 1, self.data.columns_names[i])

    def _write_data(self, ws):
        for row in range(len(self.data.data)):
            for column in range(len(self.data.columns_names)):
                _ = ws.cell(row + 2, column + 1, self.data.data[row][column])


class Write_To_Postgresql:
    """
    Insert data to selected dabatase table
    """

    def __init__(self, name, data_target_name, table, schema='public'):
        """
        :param name: Step name
        :param data_target_name: Data target name
        :param table: Selected table
        :param schema: Selcted table
        """
        self.data = None
        self.data_target_name = data_target_name
        self.logger = logging.getLogger(__name__)
        self.name = name
        self.data_targets = None
        self.conn = None
        self.host = None
        self.port = None
        self.user = None
        self.password = None
        self.schema = schema
        self.table = table

    def process(self):
        """
        Connect to database
        :return:
        """
        self.logger.info(f'Starting new read_postgresql job - {self.name}!')
        self._load_credentials()
        try:
            self.conn = psycopg2.connect(
                f'dbname={self.dbname} user={self.user} port={self.port} host={self.host} password={self.password}')
        except:
            self.logger.info(f'{self.name} - Unable to connect to database.')
            raise Exception('Unable to connect to database.')

        self._insert_data()

        self.logger.info(
            f'Read_postgresql job - {self.name} ended - {len(self.data.data)} lines, {len(self.data.columns_names)} columns.')

    def _insert_data(self):
        """
        Insert all row into table and commit it
        """
        cur = self.conn.cursor()
        for row in self.data.data:
            cur.execute(
                f'INSERT INTO {self.schema}.{self.table} ({self._prepare_header()}) VALUES ( {self._prepare_row(row)} )')

        self.conn.commit()
        cur.close()
        self.conn.close()

    def _prepare_row(self, row):
        """
        Format row in sql format
        :param row:
        :return: formatted row
        """
        output = ''
        for value in row:
            output += '\'' + str(value) + '\','
        return output[:-1]

    def _prepare_header(self):
        """
        Format header  in sql format
        :return: formatted header
        """
        return ','.join(self.data.columns_names)

    def _load_credentials(self):
        """
        Using data target name load host, dbname, port and user
        """
        config = configparser.ConfigParser()
        config.read(self.data_targets)
        if not self.data_target_name in config.sections():
            raise Exception('Invalid data target name')
        if not 'dbname' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing dbname value!')
            raise Exception('Invalid data target type')
        if not 'host' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing host value!')
            raise Exception('Invalid data target type')
        if not 'port' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing port value!')
            raise Exception('Invalid data target type')
        if not 'user' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing user value!')
            raise Exception('Invalid data target type')
        if not 'password' in config[self.data_target_name]:
            self.logger.info(f'{self.name} - missing password value!')
            raise Exception('Invalid data target type')

        self.dbname = config[self.data_target_name]['dbname']
        self.host = config[self.data_target_name]['host']
        self.port = config[self.data_target_name]['port']
        self.user = config[self.data_target_name]['user']
        self.password = config[self.data_target_name]['password']
